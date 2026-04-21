"""Variable selection and analysis module - simplified version
PSI calculation is reused in func.py, analysis.py only handles variable selection
"""
import pandas as pd
import numpy as np
import toad
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_auc_score
from joblib import Parallel, delayed


def drop_abnormal_ym(data: pd.DataFrame, min_ym_bad_sample: int = 1,
                     min_ym_sample: int = 500) -> tuple:
    """Filter abnormal months - overall statistics, not by organization"""
    stat = data.groupby('new_date_ym').agg(
        bad_cnt=('new_target', 'sum'),
        total=('new_target', 'count')
    ).reset_index()

    abnormal = stat[(stat['bad_cnt'] < min_ym_bad_sample) | (stat['total'] < min_ym_sample)]
    abnormal = abnormal.rename(columns={'new_date_ym': '年月'})
    abnormal['去除条件'] = abnormal.apply(
        lambda x: f'bad sample count {x["bad_cnt"]} less than {min_ym_bad_sample}' if x['bad_cnt'] < min_ym_bad_sample else f'total sample count {x["total"]} less than {min_ym_sample}', axis=1
    )

    if len(abnormal) > 0:
        data = data[~data['new_date_ym'].isin(abnormal['年月'])]

    # Remove empty rows
    abnormal = abnormal.dropna(how='all')
    abnormal = abnormal.reset_index(drop=True)

    return data, abnormal


def drop_highmiss_features(data: pd.DataFrame, miss_channel: pd.DataFrame,
                           threshold: float = 0.6) -> tuple:
    """Drop high missing rate features"""
    high_miss = miss_channel[miss_channel['整体缺失率'] > threshold].copy()
    high_miss['缺失率'] = high_miss['整体缺失率']

    # Modify removal condition to show specific missing rate value
    high_miss['去除条件'] = high_miss.apply(
        lambda x: f'overall missing rate is {x["缺失率"]:.4f}, exceeds threshold {threshold}', axis=1
    )

    # Remove empty rows
    high_miss = high_miss.dropna(how='all')
    high_miss = high_miss.reset_index(drop=True)

    # Drop high missing rate features
    if len(high_miss) > 0 and '变量' in high_miss.columns:
        to_drop = high_miss['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])

    return data, high_miss[['变量', '缺失率', '去除条件']]


def drop_lowiv_features(data: pd.DataFrame, features: List[str],
                       overall_iv_threshold: float = 0.05, org_iv_threshold: float = 0.02,
                       max_org_threshold: int = 8, n_jobs: int = 4) -> tuple:
    """Drop low IV features - multi-process version, returns IV details and IV processing table

    Args:
        overall_iv_threshold: Overall IV threshold, values below this are recorded in IV processing table
        org_iv_threshold: Single organization IV threshold, values below this are considered not satisfied
        max_org_threshold: Maximum tolerated organization count, if more than this number of organizations have IV below threshold, record in IV processing table

    Returns:
        data: Data after dropping
        iv_detail: IV details (IV value of each feature in each organization and overall)
        iv_process: IV processing table (features that do not meet the conditions)
    """
    from references.func import calculate_iv
    from joblib import Parallel, delayed

    orgs = sorted(data['new_org'].unique())

    print(f"   IV calculation: feature count={len(features)}, organization count={len(orgs)}")

    # Calculate IV values for all organizations at once
    def _calc_org_iv(org):
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)
        if len(org_iv) > 0:
            org_iv = org_iv.rename(columns={'IV': 'IV值'})
            org_iv['机构'] = org
            return org_iv
        return None

    # Calculate overall IV
    print(f"   Calculating overall IV...")
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    print(f"   Overall IV calculation result: {len(iv_overall)} features")
    if len(iv_overall) == 0:
        print(f"   Warning: Overall IV calculation result is empty, returning empty table")
        return data, pd.DataFrame(columns=['变量', 'IV值', '机构', '类型']), pd.DataFrame(columns=['变量', '整体IV', '低IV机构数', '处理原因'])
    iv_overall = iv_overall.rename(columns={'IV': 'IV值'})

    # Parallel calculation of IV values for all organizations
    print(f"   Parallel calculation of IV values for {len(orgs)} organizations...")
    iv_by_org_results = Parallel(n_jobs=n_jobs, verbose=0)(
        delayed(_calc_org_iv)(org) for org in orgs
    )
    iv_by_org = [r for r in iv_by_org_results if r is not None]
    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['变量', 'IV值', '机构'])
    print(f"   Organization IV summary: {len(iv_by_org)} records")

    # Convert to wide format: feature, overall, org1, org2, ..., orgn
    iv_detail_dict = {'变量': []}
    iv_detail_dict['整体'] = []

    for org in orgs:
        iv_detail_dict[org] = []

    # Get all features
    all_vars = set(iv_overall['变量'].tolist())
    if len(iv_by_org) > 0:
        all_vars.update(iv_by_org['变量'].tolist())
    all_vars = sorted(all_vars)

    for var in all_vars:
        iv_detail_dict['变量'].append(var)

        # Overall IV
        var_overall = iv_overall[iv_overall['变量'] == var]
        if len(var_overall) > 0:
            iv_detail_dict['整体'].append(var_overall['IV值'].values[0])
        else:
            iv_detail_dict['整体'].append(None)

        # IV for each organization
        for org in orgs:
            var_org = iv_by_org[iv_by_org['机构'] == org]
            var_org = var_org[var_org['变量'] == var]
            if len(var_org) > 0:
                iv_detail_dict[org].append(var_org['IV值'].values[0])
            else:
                iv_detail_dict[org].append(None)

    iv_detail = pd.DataFrame(iv_detail_dict)
    # Sort by overall IV in descending order
    iv_detail = iv_detail.sort_values('整体', ascending=False)
    iv_detail = iv_detail.reset_index(drop=True)

    # Mark features that do not meet conditions
    # 1. Overall IV below threshold
    iv_overall_low = iv_overall[iv_overall['IV值'] < overall_iv_threshold]['变量'].tolist()

    # 2. Number of organizations with single organization IV below threshold
    if len(iv_by_org) > 0:
        iv_by_org_low = iv_by_org[iv_by_org['IV值'] < org_iv_threshold].groupby('变量').size().reset_index()
        iv_by_org_low.columns = ['变量', '低IV机构数']
    else:
        iv_by_org_low = pd.DataFrame(columns=['变量', '低IV机构数'])

    # Get list of low IV organizations for each feature
    low_iv_orgs_dict = {}
    if len(iv_by_org) > 0:
        for var in iv_by_org['变量'].unique():
            var_orgs = iv_by_org[(iv_by_org['变量'] == var) & (iv_by_org['IV值'] < org_iv_threshold)]['机构'].tolist()
            low_iv_orgs_dict[var] = var_orgs

    # 3. Mark features that need processing
    iv_process = []

    # Debug info: IV distribution statistics
    if len(iv_overall) > 0:
        print(f"   Overall IV statistics: min={iv_overall['IV值'].min():.4f}, max={iv_overall['IV值'].max():.4f}, median={iv_overall['IV值'].median():.4f}")
        print(f"   Number of features with overall IV less than {overall_iv_threshold}: {(iv_overall['IV值'] < overall_iv_threshold).sum()}/{len(iv_overall)}")

    if len(iv_by_org_low) > 0:
        print(f"   Statistics of features with organization IV less than {org_iv_threshold}:")
        print(f"     Maximum low IV organization count: {iv_by_org_low['低IV机构数'].max()}")
        print(f"     Number of features with low IV organization count greater than or equal to {max_org_threshold}: {(iv_by_org_low['低IV机构数'] >= max_org_threshold).sum()}/{len(iv_by_org_low)}")

    for var in features:
        reasons = []

        # Check overall IV
        var_overall_iv = iv_overall[iv_overall['变量'] == var]['IV值'].values
        if len(var_overall_iv) > 0 and var_overall_iv[0] < overall_iv_threshold:
            reasons.append(f'overall IV {var_overall_iv[0]:.4f} less than threshold {overall_iv_threshold}')

        # Check organization IV
        var_org_low = iv_by_org_low[iv_by_org_low['变量'] == var]
        if len(var_org_low) > 0 and var_org_low['低IV机构数'].values[0] >= max_org_threshold:
            reasons.append(f'IV less than threshold {org_iv_threshold} in {var_org_low["低IV机构数"].values[0]} organizations')

        if reasons:
            iv_process.append({
                '变量': var,
                '处理原因': '; '.join(reasons),
                '低IV机构': ','.join(low_iv_orgs_dict.get(var, []))
            })

    iv_process = pd.DataFrame(iv_process)
    iv_process = iv_process.reset_index(drop=True)

    # Drop features that do not meet conditions
    if len(iv_process) > 0 and '变量' in iv_process.columns:
        to_drop = iv_process['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])

    return data, iv_detail, iv_process


def drop_highcorr_features(data: pd.DataFrame, features: List[str],
                           threshold: float = 0.8, gain_dict: dict = None, top_n_keep: int = 20) -> tuple:
    """Drop high correlation features - based on original gain, drop one feature at a time

    Args:
        data: Data
        features: Feature list
        threshold: Correlation threshold
        gain_dict: Mapping dictionary from feature to original gain
        top_n_keep: Keep top N features by original gain ranking

    Returns:
        data: Data after dropping
        dropped_info: Drop information
    """
    if gain_dict is None:
        gain_dict = {}

    # Get current feature list (only features that exist in data)
    current_features = [f for f in features if f in data.columns]

    if len(current_features) == 0:
        return data, pd.DataFrame(columns=['变量', '相关变量', '去除条件'])

    # Determine features to keep (top N by original gain)
    if gain_dict:
        # Only consider features that exist in current features
        current_gain_dict = {k: v for k, v in gain_dict.items() if k in current_features}
        if current_gain_dict:
            sorted_features = sorted(current_gain_dict.keys(), key=lambda x: current_gain_dict[x], reverse=True)
            top_features = set(sorted_features[:top_n_keep])
            # Create mapping from feature to ranking
            rank_dict = {v: i+1 for i, v in enumerate(sorted_features)}
        else:
            top_features = set()
            rank_dict = {}
    else:
        top_features = set()
        rank_dict = {}

    dropped_info = []

    # Loop to drop until no high correlation feature pairs
    while True:
        # Recalculate correlation matrix (only for current remaining features)
        current_features = [f for f in current_features if f in data.columns]
        if len(current_features) < 2:
            break

        corr = data[current_features].corr().abs()
        upper = corr.where(np.triu(np.ones(corr.shape), k=1).astype(bool))

        # Find all high correlation feature pairs
        high_corr_pairs = []
        for i, col1 in enumerate(upper.columns):
            for col2 in upper.columns[i+1:]:
                corr_val = upper.loc[col1, col2]
                if pd.notna(corr_val) and corr_val > threshold:
                    high_corr_pairs.append((col1, col2, corr_val))

        if not high_corr_pairs:
            break

        # For each high correlation feature pair, select the feature with smaller original gain as candidate for dropping
        candidates = set()
        for col1, col2, corr_val in high_corr_pairs:
            # Skip top N kept features
            if col1 in top_features and col2 in top_features:
                continue

            gain1 = gain_dict.get(col1, 0)
            gain2 = gain_dict.get(col2, 0)

            # Select feature with smaller original gain
            if gain1 <= gain2:
                candidates.add(col1)
            else:
                candidates.add(col2)

        if not candidates:
            break

        # Select feature with smallest original gain among candidates for dropping
        candidates_list = list(candidates)
        candidates_with_gain = [(c, gain_dict.get(c, 0)) for c in candidates_list]
        candidates_with_gain.sort(key=lambda x: x[1])
        to_drop = candidates_with_gain[0][0]

        # Find all features highly correlated with this feature
        related_vars = []
        for col1, col2, corr_val in high_corr_pairs:
            if col1 == to_drop:
                related_vars.append((col2, corr_val))
            elif col2 == to_drop:
                related_vars.append((col1, corr_val))

        # Record drop information
        # Related variables column: show feature name and similarity value (correlation value)
        related_str = ','.join([f"{v}(similarity={c:.4f})" for v, c in related_vars])
        # Removal condition column: show related features and their corresponding gain values
        gain_str = ','.join([f"{v}(gain={gain_dict.get(v, 0):.2f})" for v, c in related_vars])
        dropped_info.append({
            '变量': to_drop,
            '原始gain': gain_dict.get(to_drop, 0),
            '原始gain排名': rank_dict.get(to_drop, '-'),
            '相关变量': related_str,
            '去除条件': gain_str
        })

        # Delete this feature from data
        data = data.drop(columns=[to_drop], errors='ignore')
        current_features.remove(to_drop)

        print(f"   Dropped feature: {to_drop} (original gain={gain_dict.get(to_drop, 0):.2f})")

    # Convert to DataFrame and sort by original gain in descending order
    dropped_df = pd.DataFrame(dropped_info)
    if len(dropped_df) > 0:
        dropped_df = dropped_df.sort_values('原始gain', ascending=False)
        dropped_df = dropped_df.reset_index(drop=True)

    return data, dropped_df


def drop_highnoise_features(data: pd.DataFrame, features: List[str],
                           n_estimators: int = 100, max_depth: int = 5, gain_threshold: float = 50) -> tuple:
    """Null Importance to remove high noise features"""
    # Check if feature list is empty
    if len(features) == 0:
        print("   No features to process")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])

    # Check if data is sufficient
    if len(data) < 1000:
        print(f"   Insufficient data ({len(data)} rows), skip Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])

    X = data[features].copy()
    Y = data['new_target'].copy()

    # Check if X is empty or contains NaN
    if X.shape[1] == 0:
        print("   Feature data is empty, skip Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])

    # Fill NaN
    X = X.fillna(0)

    # Shuffle labels
    Y_permuted = Y.copy()
    for _ in range(20):
        Y_permuted = np.random.permutation(Y_permuted)

    clf = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )

    clf_permuted = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )

    results, results_permuted = [], []

    print("Null Importance calculation in progress...")
    for i in range(2):
        random_n = np.random.randint(30)

        X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.3, random_state=random_n)

        # Check if training data is valid
        if X_train.shape[0] == 0 or X_test.shape[0] == 0:
            print(f"  Round {i+1}: Data split failed, skip")
            continue

        clf.fit(X_train, y_train)

        X_train_, X_test_, y_train_, y_test_ = train_test_split(X, Y_permuted, test_size=0.3, random_state=random_n)

        if X_train_.shape[0] == 0 or X_test_.shape[0] == 0:
            print(f"  Round {i+1}: Shuffled data split failed, skip")
            continue

        clf_permuted.fit(X_train_, y_train_)

        imp_real = pd.DataFrame({
            'feature': clf.booster_.feature_name(),
            'gain': clf.booster_.feature_importance(importance_type='gain')
        })
        imp_perm = pd.DataFrame({
            'feature': clf_permuted.booster_.feature_name(),
            'gain': clf_permuted.booster_.feature_importance(importance_type='gain')
        })

        results.append(imp_real)
        results_permuted.append(imp_perm)

        train_auc = roc_auc_score(y_train, clf.predict_proba(X_train)[:, 1])
        test_auc = roc_auc_score(y_test, clf.predict_proba(X_test)[:, 1])
        print(f"  Round {i+1}: train_auc={train_auc:.3f}, test_auc={test_auc:.3f}")

    # Check if there are valid results
    if len(results) == 0 or len(results_permuted) == 0:
        print("   No valid training results, skip Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])

    imp_real_avg = pd.concat(results).groupby('feature')['gain'].mean().reset_index()
    imp_perm_avg = pd.concat(results_permuted).groupby('feature')['gain'].mean().reset_index()

    comparison = imp_real_avg.merge(imp_perm_avg, on='feature', suffixes=('_real', '_perm'))
    comparison['gain_real'] = comparison['gain_real'].fillna(0)
    comparison['gain_perm'] = comparison['gain_perm'].fillna(0)

    # Use condition where absolute difference of gain values before and after permutation is less than 50
    comparison['gain_diff'] = (comparison['gain_real'] - comparison['gain_perm']).abs()
    noise_features = comparison[comparison['gain_diff'] < gain_threshold]['feature'].tolist()

    # List original gain and permuted gain for all features
    dropped_info = pd.DataFrame({
        '变量': comparison['feature'].values,
        '原始gain': comparison['gain_real'].values,
        '反转后gain': comparison['gain_perm'].values
    })
    # Add status column, mark dropped features as '去除', kept features as '保留'
    dropped_info['状态'] = dropped_info.apply(
        lambda x: '去除' if np.abs(x['原始gain'] - x['反转后gain']) < gain_threshold else '保留', axis=1
    )
    # Sort by original gain in descending order
    dropped_info = dropped_info.sort_values('原始gain', ascending=False)
    dropped_info = dropped_info.reset_index(drop=True)
    # Add original gain ranking column
    dropped_info['原始gain排名'] = range(1, len(dropped_info) + 1)

    data = data.drop(columns=[c for c in noise_features if c in data.columns])

    print(f"  Dropped {len(noise_features)} noise features")
    return data, dropped_info


def _calc_single_psi(args):
    """Calculate PSI for a single organization and single feature - NaN as separate bin"""
    org, train_month, test_month, train_n, test_n, f, data_ref, min_sample = args

    try:
        org_data = data_ref[data_ref['new_org'] == org]
        train_data = org_data[org_data['new_date_ym'] == train_month]
        test_data = org_data[org_data['new_date_ym'] == test_month]

        # Get data
        train_vals = train_data[f].values
        test_vals = test_data[f].values

        # Mark NaN
        train_nan_mask = pd.isna(train_vals)
        test_nan_mask = pd.isna(test_vals)

        # Non-NaN values for binning
        train_nonan = train_vals[~train_nan_mask]
        test_nonan = test_vals[~test_nan_mask]

        if len(train_nonan) < min_sample or len(test_nonan) < min_sample:
            return {
                '机构': org, '日期': f"{train_month}->{test_month}",
                '变量': f, 'PSI': None, '有效计算': 0,
                '样本数': train_n
            }

        # Bin based on non-NaN data (10 bins)
        try:
            bins = pd.qcut(train_nonan, q=10, duplicates='drop', retbins=True)[1]
        except:
            bins = pd.cut(train_nonan, bins=10, retbins=True)[1]

        # Calculate proportion of each bin (including NaN bin)
        train_counts = []
        test_counts = []

        for i in range(len(bins)):
            if i == 0:
                train_counts.append((~train_nan_mask & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals <= bins[i])).sum())
            else:
                train_counts.append((~train_nan_mask & (train_vals > bins[i-1]) & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals > bins[i-1]) & (test_vals <= bins[i])).sum())

        # NaN bin
        train_counts.append(train_nan_mask.sum())
        test_counts.append(test_nan_mask.sum())

        # Convert to proportions
        train_pct = np.array(train_counts) / len(train_vals)
        test_pct = np.array(test_counts) / len(test_vals)

        # Avoid 0 values
        train_pct = np.where(train_pct == 0, 1e-6, train_pct)
        test_pct = np.where(test_pct == 0, 1e-6, test_pct)

        # Calculate PSI
        psi = np.sum((test_pct - train_pct) * np.log(test_pct / train_pct))

        return {
            '机构': org, '日期': f"{train_month}->{test_month}",
            '变量': f, 'PSI': round(psi, 4), '有效计算': 1,
            '样本数': train_n
        }
    except Exception as e:
        return {
            '机构': org, '日期': f"{train_month}->{test_month}",
            '变量': f, 'PSI': None, '有效计算': 0,
            '样本数': train_n
        }


def drop_highpsi_features(data: pd.DataFrame, features: List[str],
                         psi_threshold: float = 0.1, max_months_ratio: float = 1/3,
                         max_orgs: int = 4, min_sample_per_month: int = 100, n_jobs: int = 4) -> tuple:
    """Drop high PSI features - by organization + month-by-month version

    Multi-processing at feature level, loop through organizations, parallel calculation of features within organizations

    Args:
        psi_threshold: PSI threshold, values above this are considered unstable
        max_months_ratio: Maximum tolerated month ratio, if more than this ratio of months have PSI above threshold, record in processing table
        max_orgs: Maximum tolerated organization count, if more than this number of organizations are unstable, record in processing table
        min_sample_per_month: Minimum sample count per month

    Returns:
        data: Data after dropping
        psi_detail: PSI details (PSI value of each feature in each organization each month)
        psi_process: PSI processing table (features that do not meet the conditions)
    """
    orgs = data['new_org'].unique()

    # Build task list: each organization, each pair of months, each feature
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())

        if len(months) < 2:
            continue

        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]

            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]

            train_n = len(train_data)
            test_n = len(test_data)

            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample_per_month))

    # Multi-process PSI calculation (parallel at feature level)
    print(f"   PSI calculation: {len(tasks)} tasks, using {n_jobs} processes")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)

    psi_df = pd.DataFrame(results)

    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['变量', '机构', '月份', 'PSI值']), pd.DataFrame(columns=['变量', '处理原因'])

    # Filter valid calculation records
    valid_psi = psi_df[psi_df['有效计算'] == 1].copy()

    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['变量', '机构', '月份', 'PSI值']), pd.DataFrame(columns=['变量', '处理原因'])

    # PSI detail table: PSI value of each feature in each organization each month
    # Change date to single month, initial month PSI value is 0
    psi_detail = valid_psi[['机构', '日期', '变量', 'PSI']].copy()

    # Parse date, extract test month
    psi_detail['月份'] = psi_detail['日期'].apply(lambda x: x.split('->')[1] if '->' in x else x)
    psi_detail = psi_detail.rename(columns={'PSI': 'PSI值'})

    # Sort by feature, organization, month in ascending order
    psi_detail = psi_detail.sort_values(['变量', '机构', '月份'], ascending=[True, True, True])

    # Get all organizations and months
    all_orgs = sorted(psi_detail['机构'].unique())
    all_vars = sorted(psi_detail['变量'].unique())

    # Build complete PSI detail table (including initial month, PSI value is 0)
    psi_detail_list = []
    for org in all_orgs:
        org_data = psi_detail[psi_detail['机构'] == org]
        if len(org_data) == 0:
            continue

        # Get all months for this organization
        months = sorted(org_data['月份'].unique())

        for var in all_vars:
            var_data = org_data[org_data['变量'] == var]
            if len(var_data) == 0:
                continue

            # Initial month PSI value is 0
            psi_detail_list.append({
                '机构': org,
                '变量': var,
                '月份': months[0],
                'PSI值': 0.0
            })

            # Subsequent months PSI values are calculation results
            for i in range(1, len(months)):
                month = months[i]
                var_month_data = var_data[var_data['月份'] == month]
                if len(var_month_data) > 0:
                    psi_value = var_month_data['PSI值'].values[0]
                else:
                    psi_value = 0.0
                psi_detail_list.append({
                    '机构': org,
                    '变量': var,
                    '月份': month,
                    'PSI值': psi_value
                })

    psi_detail = pd.DataFrame(psi_detail_list)
    psi_detail = psi_detail[['机构', '变量', '月份', 'PSI值']]
    psi_detail = psi_detail.reset_index(drop=True)
    # Sort by feature, organization, month in ascending order
    psi_detail = psi_detail.sort_values(['变量', '机构', '月份'], ascending=[True, True, True])
    psi_detail = psi_detail.reset_index(drop=True)

    # Mark unstable
    valid_psi['不稳定'] = (valid_psi['PSI'] > psi_threshold).astype(int)

    # Summary: number of unstable months and total months for each organization each feature
    org_summary = valid_psi.groupby(['机构', '变量']).agg(
        不稳定月份数=('不稳定', 'sum'),
        总月份数=('变量', 'count')
    ).reset_index()

    # Mark whether each organization each feature is unstable
    # Ensure threshold is at least 1, avoid being too strict when organization has few months
    org_summary['不稳定阈值'] = org_summary['总月份数'].apply(
        lambda x: max(1, int(x * max_months_ratio))
    )
    org_summary['是否不稳定'] = org_summary['不稳定月份数'] >= org_summary['不稳定阈值']

    # Organization level summary: number of unstable organizations
    org_count = len(orgs)
    channel_summary = org_summary.groupby('变量').apply(
        lambda x: pd.Series({
            '机构数': org_count,
            '不稳定机构数': x['是否不稳定'].sum()
        })
    ).reset_index()

    # Mark features that need processing
    channel_summary['需处理'] = channel_summary['不稳定机构数'] >= max_orgs
    channel_summary['处理原因'] = channel_summary.apply(
        lambda x: f'PSI unstable in {x["不稳定机构数"]} organizations' if x['需处理'] else '', axis=1
    )

    # Get list of unstable organizations for each feature
    unstable_orgs_dict = {}
    for var in org_summary['变量'].unique():
        var_orgs = org_summary[(org_summary['变量'] == var) & (org_summary['是否不稳定'] == True)]['机构'].tolist()
        unstable_orgs_dict[var] = var_orgs

    # PSI processing table: features that do not meet the conditions
    psi_process = channel_summary[channel_summary['需处理']].copy()
    psi_process['不稳定机构'] = psi_process['变量'].apply(lambda x: ','.join(unstable_orgs_dict.get(x, [])))
    psi_process = psi_process[['变量', '处理原因', '不稳定机构']]
    psi_process = psi_process.reset_index(drop=True)

    # Filter features to drop
    if len(psi_process) > 0 and '变量' in psi_process.columns:
        to_drop_vars = psi_process['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])

    return data, psi_detail, psi_process


def iv_distribution_by_org(iv_detail: pd.DataFrame, oos_orgs: list = None, iv_bins: list = [0, 0.02, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """Count number and proportion of features in different IV ranges for each organization

    Args:
        iv_detail: IV detail table (containing feature, overall, organization columns)
        oos_orgs: Out-of-sample organization list
        iv_bins: IV range boundaries [0, 0.02, 0.05, 0.1, inf]

    Returns:
        IV distribution statistics table
    """
    if oos_orgs is None:
        oos_orgs = []

    # Get organization columns (exclude '变量' and '整体' columns)
    org_cols = [c for c in iv_detail.columns if c not in ['变量', '整体']]

    # Define range labels
    bin_labels = ['[0, 0.02)', '[0.02, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']

    result = []

    # Statistics for each organization (not including overall)
    for org in org_cols:
        org_iv = iv_detail[org].dropna()
        total_vars = len(org_iv)

        # Determine organization type
        org_type = '贷外' if org in oos_orgs else '建模'

        for i in range(len(iv_bins) - 1):
            lower = iv_bins[i]
            upper = iv_bins[i + 1]
            if upper == float('inf'):
                count = ((org_iv >= lower)).sum()
            else:
                count = ((org_iv >= lower) & (org_iv < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                'IV区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def psi_distribution_by_org(psi_detail: pd.DataFrame, oos_orgs: list = None, psi_bins: list = [0, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """Count number and proportion of features in different PSI ranges for each organization

    Args:
        psi_detail: PSI detail table (containing organization, feature, month, PSI value columns)
        oos_orgs: Out-of-sample organization list
        psi_bins: PSI range boundaries [0, 0.05, 0.1, inf]

    Returns:
        PSI distribution statistics table
    """
    if oos_orgs is None:
        oos_orgs = []

    # Define range labels
    bin_labels = ['[0, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']

    result = []

    # Get all organizations
    orgs = psi_detail['机构'].unique()

    for org in orgs:
        org_data = psi_detail[psi_detail['机构'] == org]

        # Determine organization type
        org_type = '贷外' if org in oos_orgs else '建模'

        # For each feature, take its maximum PSI value
        var_max_psi = org_data.groupby('变量')['PSI值'].max()
        total_vars = len(var_max_psi)

        for i in range(len(psi_bins) - 1):
            lower = psi_bins[i]
            upper = psi_bins[i + 1]
            if upper == float('inf'):
                count = ((var_max_psi >= lower)).sum()
            else:
                count = ((var_max_psi >= lower) & (var_max_psi < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                'PSI区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def value_ratio_distribution_by_org(data: pd.DataFrame, features: List[str],
                                     oos_orgs: list = None,
                                     value_bins: list = [0, 0.15, 0.35, 0.65, 0.95, 1.0]) -> pd.DataFrame:
    """Count number and proportion of features in different value ratio ranges for each organization

    Args:
        data: Data (containing new_org column)
        features: Feature list
        oos_orgs: Out-of-sample organization list
        value_bins: Value ratio range boundaries [0, 0.15, 0.35, 0.65, 0.95, 1.0]

    Returns:
        Value ratio distribution statistics table
    """
    if oos_orgs is None:
        oos_orgs = []

    # Define range labels
    bin_labels = ['[0, 15%)', '[15%, 35%)', '[35%, 65%)', '[65%, 95%)', '[95%, 100%]']

    result = []

    # Get all organizations
    orgs = data['new_org'].unique()

    for org in orgs:
        org_data = data[data['new_org'] == org]

        # Determine organization type
        org_type = '贷外' if org in oos_orgs else '建模'

        # Calculate value ratio for each feature (proportion of non-NaN)
        value_ratios = {}
        for f in features:
            if f in org_data.columns:
                non_null_count = org_data[f].notna().sum()
                total_count = len(org_data)
                value_ratios[f] = non_null_count / total_count if total_count > 0 else 0

        # Count number of features in each range
        total_vars = len(value_ratios)
        for i in range(len(value_bins) - 1):
            lower = value_bins[i]
            upper = value_bins[i + 1]
            if upper == 1.0:
                count = sum(1 for v in value_ratios.values() if lower <= v <= upper)
            else:
                count = sum(1 for v in value_ratios.values() if lower <= v < upper)
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                '有值率区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })

    return pd.DataFrame(result)


def calculate_iv_by_org(data: pd.DataFrame, features: List[str],
                        n_jobs: int = 4) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Calculate IV by organization and overall

    Returns:
        iv_by_org: IV details by organization
        iv_overall: Overall IV
    """
    from references.func import calculate_iv

    orgs = data['new_org'].unique()

    # Overall IV
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    iv_overall['类型'] = '整体'

    # IV by organization
    iv_by_org = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)  # Single process for single organization
        if len(org_iv) > 0:  # Only add non-empty results
            org_iv['机构'] = org
            org_iv['类型'] = '分机构'
            iv_by_org.append(org_iv)

    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['变量', 'IV', '机构', '类型'])

    return iv_by_org, iv_overall


def calculate_psi_detail(data: pd.DataFrame, features: List[str],
                         max_psi: float = 0.1, min_months_unstable: int = 3,
                         min_sample: int = 100, n_jobs: int = 4) -> tuple:
    """Calculate month-by-month PSI details for each feature in each organization, and mark whether to drop

    Returns:
        data: Data after dropping
        dropped: Summary of dropped features
        psi_summary: Complete PSI details (including drop flag)
    """
    orgs = data['new_org'].unique()

    # Build tasks
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())

        if len(months) < 2:
            continue

        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]

            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]

            train_n = len(train_data)
            test_n = len(test_data)

            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample))

    # Multi-process calculation
    print(f"   PSI calculation: {len(tasks)} tasks, using {n_jobs} processes")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)

    psi_df = pd.DataFrame(results)

    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '原因']), pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件'])

    # Filter valid calculation records
    valid_psi = psi_df[psi_df['有效计算'] == 1].copy()

    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '原因']), pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件'])

    # Mark unstable
    valid_psi['不稳定'] = (valid_psi['PSI'] > max_psi).astype(int)

    # Summary: number of unstable months for each organization each feature
    org_summary = valid_psi.groupby(['机构', '变量'])['不稳定'].sum().reset_index()
    org_summary.columns = ['机构', '变量', '不稳定月份数']

    # Organization level summary: features with more than min_months_unstable unstable months
    org_count = len(orgs)
    channel_summary = org_summary.groupby('变量').apply(
        lambda x: pd.Series({
            '机构数': org_count,
            '不稳定机构数': (x['不稳定月份数'] >= min_months_unstable).sum()
        })
    ).reset_index()

    # Mark features that need to be dropped (more than 1/3 organizations unstable)
    channel_summary['需剔除'] = channel_summary['不稳定机构数'] > (channel_summary['机构数'] / 3)
    channel_summary['是否剔除'] = channel_summary['需剔除'].astype(int)
    channel_summary['去除条件'] = channel_summary.apply(
        lambda x: f'More than 1/3 of {org_count} organizations have PSI>{max_psi} for {min_months_unstable} consecutive months' if x['需剔除'] else '', axis=1
    )

    # Filter features to drop
    if len(channel_summary) > 0 and '变量' in channel_summary.columns:
        to_drop_vars = channel_summary[channel_summary['需剔除']]['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])

    # Organize drop information (only return dropped features)
    dropped = channel_summary[channel_summary['需剔除']].copy()
    dropped['原因'] = f'More than 1/3 of {org_count} organizations have PSI>{max_psi} for {min_months_unstable} consecutive months'

    return data, dropped[['变量', '机构数', '不稳定机构数', '原因']], channel_summary[['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件']]


def export_cleaning_report(filepath: str, steps: list,
                           iv_detail: pd.DataFrame = None,
                           iv_process: pd.DataFrame = None,
                           psi_detail: pd.DataFrame = None,
                           psi_process: pd.DataFrame = None,
                           params: dict = None,
                           iv_distribution: pd.DataFrame = None,
                           psi_distribution: pd.DataFrame = None,
                           value_ratio_distribution: pd.DataFrame = None):
    """Export cleaning report to xlsx - one sheet per step

    Args:
        filepath: Output path
        steps: Cleaning step list [(step name, DataFrame), ...]
        iv_detail: IV details (IV value of each feature in each organization and overall)
        iv_process: IV processing table (features that do not meet the conditions)
        psi_detail: PSI details (PSI value of each feature in each organization each month)
        psi_process: PSI processing table (features that do not meet the conditions)
        params: Hyperparameter dictionary, used to dynamically generate conditions
        iv_distribution: IV distribution statistics table
        psi_distribution: PSI distribution statistics table
        value_ratio_distribution: Value ratio distribution statistics table
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(filepath)
    except:
        wb = Workbook()
        wb.remove(wb.active)

    # Summary sheet - only show real filtering steps
    if '汇总' in wb.sheetnames:
        del wb['汇总']
    ws = wb.create_sheet('汇总', 0)
    ws['A1'] = 'Data Cleaning Report'
    ws['A2'] = f'Generated at: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A4'] = 'Step'
    ws['B4'] = 'Operation Details'
    ws['C4'] = 'Operation Result'
    ws['D4'] = 'Condition'

    # Only show real filtering steps (excluding details and distribution statistics)
    filter_steps = [
        'Step4-异常月份处理', 'Step6-高缺失率处理', 'Step7-IV处理',
        'Step8-PSI处理', 'Step9-null importance处理', 'Step10-高相关性剔除'
    ]

    # Steps to exclude (details and distribution statistics)
    exclude_steps = [
        'Step7-IV明细', 'Step7-IV分布统计', 'Step8-PSI明细',
        'Step8-PSI分布统计', 'Step5-有值率分布统计'
    ]

    # Steps that need to show drop count
    show_drop_count_steps = ['分离OOS数据']

    # Steps that only show parameter standards (no operation result)
    show_param_only_steps = ['机构样本统计', '缺失率明细']

    # Add note: each step is executed independently
    ws['A3'] = 'Note: Each filtering step is executed independently, data is not deleted, only statistics of features that do not meet conditions are recorded'

    # Get parameters, use default values if not provided
    if params is None:
        params = {}

    min_ym_bad_sample = params.get('min_ym_bad_sample', 10)
    min_ym_sample = params.get('min_ym_sample', 500)
    missing_ratio = params.get('missing_ratio', 0.6)
    overall_iv_threshold = params.get('overall_iv_threshold', 0.1)
    org_iv_threshold = params.get('org_iv_threshold', 0.1)
    max_org_threshold = params.get('max_org_threshold', 2)
    psi_threshold = params.get('psi_threshold', 0.1)
    max_months_ratio = params.get('max_months_ratio', 1/3)
    max_orgs = params.get('max_orgs', 4)
    gain_threshold = params.get('gain_threshold', 50)

    step_num = 1
    for name, df in steps:
        # Skip detail and distribution statistics steps
        if name in exclude_steps:
            continue

        # Remove StepX- prefix from operation details
        display_name = name.replace('Step4-', '').replace('Step6-', '').replace('Step7-', '').replace('Step8-', '').replace('Step9-', '').replace('Step10-', '')

        # Only show parameter standard steps (no operation result)
        if name in show_param_only_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            result = ''
            # Condition: show parameter standards
            if name == '机构样本统计':
                condition = 'Statistics of sample count and bad sample rate for each organization'
            elif name == '缺失率明细':
                condition = 'Calculate missing rate for each feature'
            else:
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        # Show steps that need to display drop count
        elif name in show_drop_count_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            if df is not None and len(df) > 0:
                if name == '分离OOS数据':
                    # Special handling: show OOS and modeling sample counts
                    if '变量' in df.columns and '数量' in df.columns:

                        oos_count = df[df['变量'] == 'OOS样本']['数量'].values[0] if len(df[df['变量'] == 'OOS样本']) > 0 else 0
                        model_count = df[df['变量'] == '建模样本']['数量'].values[0] if len(df[df['变量'] == '建模样本']) > 0 else 0
                        result = f'OOS samples {oos_count}, modeling samples {model_count}'
                    else:
                        result = f'{len(df)} rows'
                elif '变量' in df.columns:
                    result = f'Dropped {len(df)} features'
                else:
                    result = f'Dropped {len(df)}'
                condition = ''
            else:
                result = 'Empty'
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        elif name in filter_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)

            # Generate operation result and condition
            if df is not None and len(df) > 0:
                if name == 'Step4-异常月份处理':
                    # Operation result: dropped months
                    if '年月' in df.columns:
                        result = 'Dropped ' + ','.join(df['年月'].astype(str).tolist())
                    else:
                        result = 'Dropped ' + ','.join(df.iloc[:, 0].astype(str).tolist())
                    # Condition: parameter standards
                    condition = f'Months with bad sample count less than {min_ym_bad_sample} or total sample count less than {min_ym_sample} will be dropped (independent execution)'
                elif name == 'Step6-高缺失率处理':
                    # Operation result: number of dropped features
                    if '变量' in df.columns:
                        result = f'Dropped {len(df)} features'
                    else:
                        result = f'Dropped {len(df)}'
                    # Condition: parameter standards
                    condition = f'Features with overall missing rate greater than {missing_ratio} will be dropped (independent execution)'
                elif name == 'Step7-IV处理':
                    # Operation result: number of dropped features
                    if '变量' in df.columns:
                        result = f'Dropped {len(df)} features'
                    else:
                        result = f'Dropped {len(df)}'
                    # Condition: parameter standards
                    condition = f'Features with overall IV less than {overall_iv_threshold} or IV less than {org_iv_threshold} in {max_org_threshold} or more organizations will be dropped (independent execution)'
                elif name == 'Step8-PSI处理':
                    # Operation result: number of dropped features
                    if '变量' in df.columns:
                        result = f'Dropped {len(df)} features'
                    else:
                        result = f'Dropped {len(df)}'
                    # Condition: parameter standards
                    condition = f'PSI threshold {psi_threshold}, if an organization has more than {max_months_ratio:.0%} months with PSI greater than {psi_threshold}, the organization is considered unstable, if more than {max_orgs} organizations are unstable, the feature will be dropped (independent execution)'
                elif name == 'Step9-null importance处理':
                    # Operation result: number of dropped features
                    if '变量' in df.columns:
                        result = f'Dropped {len(df[df["状态"]=="去除"])} features'
                    else:
                        result = f'Dropped {len(df)}'
                    # Condition: parameter standards
                    condition = f'Features with absolute difference of gain values before and after permutation less than {gain_threshold} will be identified as noise and dropped (independent execution)'
                elif name == 'Step10-高相关性剔除':
                    # Operation result: number of dropped features
                    if '变量' in df.columns:
                        result = f'Dropped {len(df)} features'
                    else:
                        result = f'Dropped {len(df)}'
                    # Condition: parameter standards
                    max_corr = params.get('max_corr', 0.9)
                    top_n_keep = params.get('top_n_keep', 20)
                    condition = f'Features with correlation greater than {max_corr} will be dropped, keep top {top_n_keep} features by original gain ranking (independent execution)'
                else:
                    result = 'Dropped ' + str(len(df))
                    condition = ''
            else:
                result = 'Empty'
                condition = ''

            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1

    # Calculate total number of dropped features (take union of dropped features from each step)
    all_dropped_vars = set()
    for name, df in steps:
        if name in filter_steps and df is not None and len(df) > 0 and '变量' in df.columns:
            if name == 'Step9-null importance处理':
                # null importance processing needs to filter features with status "去除"
                dropped_vars = df[df['状态'] == '去除']['变量'].tolist()
            else:
                dropped_vars = df['变量'].tolist()
            # Take union (deduplicated)
            all_dropped_vars = all_dropped_vars.union(set(dropped_vars))

    # Add final statistics row
    final_step_num = step_num
    ws.cell(4+final_step_num, 1, final_step_num)
    ws.cell(4+final_step_num, 2, 'Final Dropped Features Statistics')
    ws.cell(4+final_step_num, 3, f'Total dropped {len(all_dropped_vars)} features (union of dropped features from each step)')
    ws.cell(4+final_step_num, 4, 'Each step is executed independently, final dropped features are the union of dropped features from each step')

    # Details of each step (create sheets in step progression order)
    # Define sheet creation order
    sheet_order = [
        '机构样本统计', '分离OOS数据', 'Step4-异常月份处理', '缺失率明细',
        'Step5-有值率分布统计', 'Step6-高缺失率处理', 'Step7-IV明细', 'Step7-IV处理',
        'Step7-IV分布统计', 'Step8-PSI明细', 'Step8-PSI处理', 'Step8-PSI分布统计',
        'Step9-null importance处理', 'Step10-高相关性剔除'
    ]

    # Create sheets in order
    for sheet_name in sheet_order:
        # Find corresponding DataFrame in steps
        df = None
        for name, step_df in steps:
            if name == sheet_name:
                df = step_df
                break

        if df is not None:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_detail = wb.create_sheet(sheet_name)

            for j, col in enumerate(df.columns):
                ws_detail.cell(1, j+1, col)

            for i, row in df.iterrows():
                for j, val in enumerate(row):
                    # Write value directly, avoid character escaping issues
                    ws_detail.cell(i+2, j+1, val if val is not None else '')

            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font

    wb.save(filepath)
    print(f"Report saved: {filepath}")